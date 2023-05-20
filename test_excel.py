from openpyxl import Workbook
wb = Workbook()


# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

#ワークシートを追加
ws1 = wb.create_sheet(title="Mysheet") 
data = 0
for x in range(1,11):
    for y in range(1,11):
        date = 10 * (x - 1) + y
        ws1.cell(row=x, column=y).value = date

#シート名の確認
print( wb.sheetnames )

# Save the file
wb.save("sample.xlsx")
